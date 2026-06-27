"""
日本年金機構 適用事業所検索 から被保険者数を取得するスクリプト
https://www.nenkin.go.jp/service/kounen/jigyosho/jigyoshokensaku.html

【使い方】
1. 依存パッケージをインストール:
   pip install playwright openpyxl
   playwright install chromium

2. 取得したいExcelファイルを指定して実行:
   python fetch_insured_count_from_nenkin.py 営業リスト_被保険者数追加.xlsx

3. 結果はExcelのK列（被保険者数）とL列（情報源）に書き込まれます
"""

import sys
import re
import time
import json
import openpyxl
from pathlib import Path
from playwright.sync_api import sync_playwright


SEARCH_URL = "https://www.nenkin.go.jp/do/search_section/"


def clean_company_name(name: str) -> str:
    """企業名をクリーンアップして検索しやすくする"""
    if not name:
        return ""
    # 全角スペースや空白を除去
    name = re.sub(r"[\s　]+", "", name)
    # 法人格の表記ゆれを正規化
    name = name.replace("（株）", "株式会社").replace("(株)", "株式会社")
    name = name.replace("（有）", "有限会社").replace("(有)", "有限会社")
    # OCRノイズを除去（記号や英数字の連続）
    name = re.sub(r"[A-Z]{3,}[\s/\-].*", "", name)  # 末尾のOCRゴミ
    return name.strip()


def search_company(page, company_name: str) -> dict | None:
    """
    適用事業所検索で会社を検索して被保険者数を取得
    Returns: {"insured_count": int, "source": str} or None
    """
    try:
        page.goto(SEARCH_URL, timeout=30000, wait_until="networkidle")

        # 事業所名入力フィールドを探す
        # ページによってフィールド名が異なる可能性があるため複数パターン試行
        name_selectors = [
            'input[name="jigyoshoName"]',
            'input[id*="jigyoshoName"]',
            'input[name*="name"]',
            'input[placeholder*="事業所"]',
            'input[placeholder*="名称"]',
            'input[type="text"]:first-of-type',
        ]

        name_input = None
        for sel in name_selectors:
            try:
                elem = page.wait_for_selector(sel, timeout=3000)
                if elem:
                    name_input = elem
                    break
            except Exception:
                continue

        if not name_input:
            print(f"  [警告] 入力フィールドが見つかりません: {company_name}")
            return None

        # 企業名を入力してEnterまたは検索ボタンをクリック
        name_input.clear()
        name_input.fill(company_name)
        time.sleep(0.5)

        # 検索ボタンをクリック
        search_btn_selectors = [
            'button[type="submit"]',
            'input[type="submit"]',
            'button:has-text("検索")',
            'a:has-text("検索")',
        ]
        for sel in search_btn_selectors:
            try:
                btn = page.query_selector(sel)
                if btn:
                    btn.click()
                    break
            except Exception:
                continue
        else:
            page.keyboard.press("Enter")

        page.wait_for_load_state("networkidle", timeout=15000)
        time.sleep(1)

        # 結果から被保険者数を抽出
        return extract_insured_count(page, company_name)

    except Exception as e:
        print(f"  [エラー] {company_name}: {e}")
        return None


def extract_insured_count(page, company_name: str) -> dict | None:
    """検索結果ページから被保険者数を抽出"""
    content = page.content()

    # 被保険者数パターンを探す
    patterns = [
        r"被保険者数[：:\s]*(\d+)",
        r"(\d+)\s*名",
        r"(\d+)\s*人",
    ]

    # テーブルの行から会社名と被保険者数を探す
    rows = page.query_selector_all("tr, .result-row, .search-result")
    for row in rows:
        text = row.inner_text()
        if not text.strip():
            continue

        # 企業名の一部が含まれるか確認
        clean_name = re.sub(r"[株式有限合同会社\s（）()]", "", company_name)
        if len(clean_name) > 3 and clean_name not in text:
            continue

        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                count = int(match.group(1))
                if 1 <= count <= 999999:  # 妥当な範囲
                    return {
                        "insured_count": count,
                        "source": f"nenkin.go.jp 適用事業所検索 ({company_name})"
                    }

    # 直接ページのテキストから探す
    full_text = page.inner_text("body")
    for pattern in patterns:
        matches = re.findall(pattern, full_text)
        for m in matches:
            count = int(m)
            if 1 <= count <= 999999:
                return {
                    "insured_count": count,
                    "source": f"nenkin.go.jp 適用事業所検索 ({company_name})"
                }

    return None


def process_excel(excel_path: str):
    """Excelファイルを読み込んで被保険者数を取得・記入"""
    path = Path(excel_path)
    if not path.exists():
        print(f"ファイルが見つかりません: {excel_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # ヘッダー行を確認
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"列: {headers}")

    # 被保険者数列を探す（なければK列=11列目に追加）
    insured_col = None
    source_col = None
    for i, h in enumerate(headers, 1):
        if h and "被保険者数" in str(h):
            insured_col = i
        if h and "情報源" in str(h) and insured_col:
            source_col = i

    if not insured_col:
        insured_col = len(headers) + 1
        ws.cell(1, insured_col).value = "被保険者数"
    if not source_col:
        source_col = insured_col + 1
        ws.cell(1, source_col).value = "被保険者数情報源"

    # 処理対象行（被保険者数が空の行のみ）を収集
    targets = []
    name_col = 3  # C列（企業名）
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, name_col).value
        existing = ws.cell(row, insured_col).value
        if name and not existing:
            clean = clean_company_name(str(name))
            if clean:
                targets.append({"row": row, "name": str(name), "clean_name": clean})

    print(f"処理対象: {len(targets)}社（被保険者数未取得）")

    results = {"found": 0, "not_found": 0, "errors": 0}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
            locale="ja-JP",
        )
        page = context.new_page()

        for i, target in enumerate(targets):
            print(f"[{i+1}/{len(targets)}] {target['clean_name']}...", end=" ", flush=True)

            result = search_company(page, target["clean_name"])

            if result:
                ws.cell(target["row"], insured_col).value = result["insured_count"]
                ws.cell(target["row"], source_col).value = result["source"]
                print(f"✓ {result['insured_count']}名")
                results["found"] += 1
            else:
                print("- 見つからず")
                results["not_found"] += 1

            # 3社ごとに保存
            if (i + 1) % 3 == 0:
                wb.save(excel_path)

            # レート制限対策（サーバー負荷軽減）
            time.sleep(1.5)

        browser.close()

    wb.save(excel_path)

    print(f"\n完了！")
    print(f"  取得成功: {results['found']}社")
    print(f"  見つからず: {results['not_found']}社")
    print(f"  保存先: {excel_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python fetch_insured_count_from_nenkin.py <Excelファイルパス>")
        print("例:     python fetch_insured_count_from_nenkin.py 営業リスト_被保険者数追加.xlsx")
        sys.exit(1)

    process_excel(sys.argv[1])
