"""
Google Sheetsのデータを Excel / CSV としてエクスポートする。
openpyxlのみ使用（xlsxwriter不要）。
"""
import csv
import logging
from datetime import datetime
from pathlib import Path

from ..config import load_config
from ..sheets.client import SheetsClient, SALES_COLUMNS

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path("output")


def _make_filename(category: str, fmt: str) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    cat_label = category or "all"
    return OUTPUT_DIR / f"sales_list_{cat_label}_{ts}.{fmt}"


def export(
    fmt: str = "xlsx",
    category_filter: str | None = None,
    config_path: str | None = None,
) -> Path:
    """
    Google Sheetsの「営業リスト」を Excel または CSV でエクスポートする。
    送信可否=不可 の行は含めない。
    戻り値: 出力ファイルパス
    """
    config = load_config(config_path)
    sheets_cfg  = config.get("google_sheets", {})
    sales_sheet = sheets_cfg.get("sales_sheet_name", "営業リスト")

    sheets = SheetsClient(
        spreadsheet_id=sheets_cfg["spreadsheet_id"],
        credentials_dir="credentials",
    )

    rows = sheets.get_all_rows(sales_sheet)
    if not rows:
        raise RuntimeError("営業リストシートが空です")

    header = rows[0]
    col_map = {h: i for i, h in enumerate(header)}

    def cell(row, col):
        i = col_map.get(col, -1)
        return row[i] if i >= 0 and i < len(row) else ""

    # フィルタリング: 送信可否=可 のみ、カテゴリ指定があれば絞り込み
    filtered = []
    for row in rows[1:]:
        if cell(row, "送信可否") != "可":
            continue
        if category_filter and cell(row, "対象カテゴリ") != category_filter:
            continue
        filtered.append(row)

    logger.info(f"エクスポート対象: {len(filtered)}件")

    OUTPUT_DIR.mkdir(exist_ok=True)
    out_path = _make_filename(category_filter or "all", fmt)

    if fmt == "csv":
        _write_csv(out_path, header, filtered)
    else:
        _write_xlsx(out_path, header, filtered)

    logger.info(f"出力完了: {out_path}")
    return out_path


def _write_csv(path: Path, header: list, rows: list[list]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in rows:
            # 列数を揃える
            padded = row + [""] * max(0, len(header) - len(row))
            writer.writerow(padded[:len(header)])


def _write_xlsx(path: Path, header: list, rows: list[list]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxlが必要です: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "営業リスト"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ヘッダ行
    for col_idx, col_name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align

    # データ行
    wrap_cols = {"メール本文", "フォーム送信用文面", "提案仮説",
                 "SNS上の改善余地", "具体的に褒めるべきポイント",
                 "月50万円以上を払える可能性がある理由", "法人と判断した根拠"}

    for row_idx, row in enumerate(rows, start=2):
        padded = row + [""] * max(0, len(header) - len(row))
        for col_idx, value in enumerate(padded[:len(header)], start=1):
            col_name = header[col_idx - 1]
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name in wrap_cols:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # 列幅の自動調整（最大50文字）
    for col_idx, col_name in enumerate(header, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in wrap_cols:
            ws.column_dimensions[col_letter].width = 40
        else:
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(rows)+2)),
                default=10
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"
    wb.save(path)
